import pandas as pd
import serial
import time
import openpyxl
import os
from openpyxl import load_workbook
from datetime import datetime

clientes_file = os.path.join(os.getcwd(), 'clientes.xlsx')
registros_file = os.path.join(os.getcwd(), 'registros.xlsx')

sheet_name = 'Clientes' 
registro_sheet = 'Registros' 

def recargar_datos():
    global df
    try:
        df = pd.read_excel(clientes_file, sheet_name=sheet_name, engine='openpyxl')
    except Exception as e:
        print(f"Error al recargar datos desde el archivo Excel: {e}")
        exit()

def agregar_cliente_excel(nuevo_cliente):
    try:
        book = load_workbook(clientes_file)
        sheet = book[sheet_name]

        for row in nuevo_cliente.itertuples(index=False, name=None):
            sheet.append(row)

        book.save(clientes_file)
        print("Cliente agregado exitosamente.")
    except Exception as e:
        print(f"Error al agregar cliente al archivo Excel: {e}")
        exit()

def eliminar_cliente_excel(uid):
    global df
    try:
        df = pd.read_excel(clientes_file, sheet_name=sheet_name, engine='openpyxl')
        
        df = df[df['UID'] != uid]
        
        df.to_excel(clientes_file, sheet_name=sheet_name, index=False, engine='openpyxl')
        print(f"Cliente con UID {uid} eliminado exitosamente.")
    except Exception as e:
        print(f"Error al eliminar cliente del archivo Excel: {e}")
        exit()

def registrar_acceso(uid, nombre, resultado):
    try:
        if not os.path.exists(registros_file):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = registro_sheet
            sheet.append(["Fecha y Hora", "UID", "Nombre", "Resultado"])
            book.save(registros_file)
        else:
            book = load_workbook(registros_file)
            sheet = book[registro_sheet]

        fecha_hora = datetime.now().strftime("%Y-%m-%d - %H:%M:%S")
        sheet.append([fecha_hora, uid, nombre, resultado])

        book.save(registros_file)
        print("Acceso registrado exitosamente.")
    except Exception as e:
        print(f"Error al registrar acceso: {e}")

recargar_datos()

try:
    ser = serial.Serial('COM3', 9600, timeout=1)
    time.sleep(2)
except serial.SerialException as e:
    print(f"Error al conectar con el puerto serie: {e}")
    exit()

master_uid = "5363f0757101"

def buscar_nombre_por_uid(uid):
    cliente = df[df['UID'] == uid]
    if not cliente.empty:
        return cliente.iloc[0]['Nombre']
    return None

try:
    while True:
        if ser.in_waiting > 0:
            uid = ser.readline().decode('utf-8').strip()

            if "Alta UID:" in uid:
                nuevo_uid = uid.split(":")[1].strip()
                print(f"Nuevo UID detectado: {nuevo_uid}")

                if buscar_nombre_por_uid(nuevo_uid) is not None:
                    ser.write(f"UID ya registrado: {nuevo_uid}\n".encode('utf-8'))
                    print(f"Error: El UID {nuevo_uid} ya está registrado.")
                else:
                    nuevo_nombre = input("Ingrese el nombre del nuevo usuario: ")

                    nuevo_cliente = pd.DataFrame({'UID': [nuevo_uid], 'Nombre': [nuevo_nombre]})
                    agregar_cliente_excel(nuevo_cliente)

                    ser.write(f"Nuevo UID registrado: {nuevo_uid}\n".encode('utf-8'))
                    print(f"Nuevo UID registrado: {nuevo_uid} con el nombre: {nuevo_nombre}")
                    
                    recargar_datos()

            elif "Baja UID:" in uid:
                uid_a_borrar = uid.split(":")[1].strip()
                print(f"Solicitud de eliminación de UID: {uid_a_borrar}")

                if buscar_nombre_por_uid(uid_a_borrar) is None:
                    ser.write(f"UID no encontrado: {uid_a_borrar}\n".encode('utf-8'))
                    print(f"Error: El UID {uid_a_borrar} no está registrado.")
                else:
                    eliminar_cliente_excel(uid_a_borrar)

                    ser.write(f"UID eliminado: {uid_a_borrar}\n".encode('utf-8'))
                    print(f"UID eliminado: {uid_a_borrar}")
                    
                    recargar_datos()

            elif uid == master_uid:
                ser.write(f"Acceso permitido\n".encode('utf-8'))
                print("Acceso permitido al UID maestro")
                registrar_acceso(uid, "Maestro", "Acceso permitido")
            else:
                recargar_datos()

                nombre = buscar_nombre_por_uid(uid)
                if nombre is not None:
                    ser.write(f"Acceso permitido a: {nombre}\n".encode('utf-8'))
                    print(f"Acceso permitido a: {nombre}")
                    registrar_acceso(uid, nombre, "Acceso permitido")
                else:
                    ser.write("Acceso denegado\n".encode('utf-8'))
                    print(f"Acceso denegado para UID: {uid}")
                    registrar_acceso(uid, "Desconocido", "Acceso denegado")

except KeyboardInterrupt:
    print("Script interrumpido.")
finally:
    ser.close()
    print("Puerto serie cerrado.")
