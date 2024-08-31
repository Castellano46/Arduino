import pandas as pd
import serial
import time
import os
import openpyxl 
from openpyxl import load_workbook
from datetime import datetime, timedelta

clientes_file = os.path.join(os.getcwd(), 'clientes.xlsx')
registros_file = os.path.join(os.getcwd(), 'registros.xlsx')
sheet_name = 'Clientes'
registro_sheet = 'Registros'

master_uid = "5363f0757101"

def recargar_datos():
    try:
        df = pd.read_excel(clientes_file, sheet_name=sheet_name, engine='openpyxl')
        return df
    except Exception as e:
        print(f"Error al recargar datos desde el archivo Excel: {e}")
        return pd.DataFrame()

def agregar_cliente_excel(nuevo_cliente):
    try:
        book = load_workbook(clientes_file)
        sheet = book[sheet_name]
        for row in nuevo_cliente.itertuples(index=False, name=None):
            sheet.append(row)
        book.save(clientes_file)
    except Exception as e:
        print(f"Error al agregar cliente al archivo Excel: {e}")

def eliminar_cliente_excel(uid):
    try:
        if uid == master_uid:
            print(f"Error: No se puede eliminar el UID maestro ({master_uid}).")
            return "UID maestro, no se puede eliminar"
        
        df = pd.read_excel(clientes_file, sheet_name=sheet_name, engine='openpyxl')
        df = df[df['UID'] != uid]
        df.to_excel(clientes_file, sheet_name=sheet_name, index=False, engine='openpyxl')
        return "UID eliminado"
    except Exception as e:
        print(f"Error al eliminar cliente del archivo Excel: {e}")
        return "Error al eliminar el cliente"

def registrar_acceso(uid, nombre, resultado, estado):
    try:
        if not os.path.exists(registros_file):
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = registro_sheet
            sheet.append(["Fecha y Hora", "UID", "Nombre", "Resultado", "Estado"])
            book.save(registros_file)
        else:
            book = load_workbook(registros_file)
            sheet = book[registro_sheet]

        fecha_hora = datetime.now().strftime("%Y-%m-%d - %H:%M:%S")
        sheet.append([fecha_hora, uid, nombre, resultado, estado])
        book.save(registros_file)
        print("Acceso registrado exitosamente.")
    except Exception as e:
        print(f"Error al registrar acceso: {e}")

def determinar_estado(uid):
    if not os.path.exists(registros_file):
        return "Entrada"

    book = load_workbook(registros_file)
    sheet = book[registro_sheet]
    registros = sheet.iter_rows(values_only=True)
    next(registros)

    ahora = datetime.now()
    ultimo_estado = None

    for row in reversed(list(registros)):
        if row[1] == uid:
            ultima_vez = datetime.strptime(row[0], "%Y-%m-%d - %H:%M:%S")
            diferencia = ahora - ultima_vez
            ultimo_estado = row[4]

            if diferencia <= timedelta(hours=10):
                if ultimo_estado == "Entrada":
                    return "Salida"
                else:
                    return "Entrada"
            else:
                break

    return "Entrada"

def buscar_nombre_por_uid(uid):
    df = recargar_datos()
    cliente = df[df['UID'] == uid]
    if not cliente.empty:
        return cliente.iloc[0]['Nombre']
    return None

def iniciar_serial():
    try:
        ser = serial.Serial('COM3', 9600, timeout=1)
        time.sleep(2)
        return ser
    except serial.SerialException as e:
        print(f"Error al conectar con el puerto serie: {e}")
        return None

def manejar_acceso():
    ser = iniciar_serial()
    try:
        while True:
            if ser and ser.in_waiting > 0:
                uid = ser.readline().decode('utf-8').strip()

                if "Alta UID:" in uid:
                    nuevo_uid = uid.split(":")[1].strip()
                    print(f"Nuevo UID detectado: {nuevo_uid}")

                    if buscar_nombre_por_uid(nuevo_uid) is not None:
                        ser.write(f"UID ya registrado: {nuevo_uid}\n".encode('utf-8'))
                        print(f"Error: El UID {nuevo_uid} ya está registrado.")
                    else:
                        nuevo_nombre = input("Ingrese el nombre del nuevo usuario: ")

                        nuevo_cliente = pd.DataFrame({'UID': [nuevo_uid], 'Nombre': [nuevo_nombre]})
                        agregar_cliente_excel(nuevo_cliente)

                        ser.write(f"Nuevo UID registrado: {nuevo_uid}\n".encode('utf-8'))
                        print(f"Nuevo UID registrado: {nuevo_uid} con el nombre: {nuevo_nombre}")
                        recargar_datos()

                elif "Baja UID:" in uid:
                    uid_a_borrar = uid.split(":")[1].strip()
                    print(f"Solicitud de eliminación de UID: {uid_a_borrar}")

                    resultado = eliminar_cliente_excel(uid_a_borrar)
                    ser.write(f"{resultado}\n".encode('utf-8'))

                    if resultado == "UID maestro, no se puede eliminar":
                        print(f"Intento de eliminar el UID maestro {uid_a_borrar} prevenido.")
                    else:
                        recargar_datos()

                elif uid == master_uid:
                    ser.write(f"Acceso permitido\n".encode('utf-8'))
                    print("Acceso permitido a Mari")
                    registrar_acceso(uid, "Mari", "Acceso permitido", "Entrada")
                else:
                    recargar_datos()

                    nombre = buscar_nombre_por_uid(uid)
                    if nombre is not None:
                        estado = determinar_estado(uid)
                        ser.write(f"Acceso permitido a: {nombre}\n".encode('utf-8'))
                        print(f"Acceso permitido a: {nombre} ({estado})")
                        registrar_acceso(uid, nombre, "Acceso permitido", estado)
                    else:
                        ser.write("Acceso denegado\n".encode('utf-8'))
                        print(f"Acceso denegado para UID: {uid}")
                        registrar_acceso(uid, "Desconocido", "Acceso denegado", "Desconocido")

    except KeyboardInterrupt:
        print("Script interrumpido.")
    finally:
        if ser:
            ser.close()
            print("Puerto serie cerrado.")
